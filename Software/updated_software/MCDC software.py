import sys
import glob
from PyQt5.QtWidgets import QApplication, QFileDialog, QMainWindow, QWidget, QMessageBox
from PyQt5 import QtCore, uic
import truthTableGenerator
import time
from PyQt5.QtCore import pyqtSignal, QThread
import openpyxl
from functools import partial


class MCDC_PROGRAM(QtCore.QObject):
    progressChanged = QtCore.pyqtSignal(int)
    maximumChanged = QtCore.pyqtSignal(int)
    resultChanged = QtCore.pyqtSignal(str)

    @QtCore.pyqtSlot(str, str)
    def CreateMCDC(self, inputFile, outputFile):
        path = inputFile
        wb = openpyxl.load_workbook(path)
        self.sheet_obj = wb.active

        self.No_of_Variables = self.sheet_obj.max_column-2

        self.No_of_combination = 2**self.No_of_Variables

        expression = self.sheet_obj.cell(row=1, column=2).value
        self.selectedRowList = []

        progress = 1

        self.maximumChanged.emit(self.No_of_combination*self.No_of_Variables)
        # loop to check for every variable and incremented by 2 to match variable column
        for i in range(2, self.No_of_Variables+2):
            # creating 2 different lists to store T and F value of every variable
            exec("%s  = []" % ("self.VT" + str(i-1)))
            exec("%s  = []" % ("self.VF" + str(i-1)))
            # To check for every combination and incremented by 4 to match row number
            for j in range(4, self.No_of_combination+4):
                self.progressChanged.emit(progress)
                for k in range(2, self.No_of_Variables+2):

                    exec("%s  = %d" % ((self.sheet_obj.cell(row=3, column=k).value).strip(),
                                       self.converted((self.sheet_obj.cell(row=j, column=k).value).strip())))

                exec("%s  = %d" % ((self.sheet_obj.cell(row=3, column=i).value).strip(),
                                   self.Reverse((self.sheet_obj.cell(row=j, column=i).value).strip())))

                try:
                    new_result = eval(expression)

                    if (new_result != self.converted((self.sheet_obj.cell(row=j, column=self.sheet_obj.max_column).value).strip())):

                        if (self.converted((self.sheet_obj.cell(row=j, column=self.sheet_obj.max_column).value).strip()) == 1):
                            eval("self.VT"+str(i-1)).append(j)
                        else:
                            eval("self.VF"+str(i-1)).append(j)

                except:
                    # return "expression"
                    return self.resultChanged.emit("expression")
                progress += 1

        # Genetrate final list
        for w in range(1, self.No_of_Variables+1):

            F = self.makeFinalListFalse("self.VF"+str(w), w)

            if F != None:
                if F not in self.selectedRowList:
                    self.selectedRowList.append(F)
                self.makeFinalListTrue(F, w)

        # print(self.selectedRowList)
        # Storing required rows in different lists

        R1 = []
        R2 = []
        R3 = []

        for v in range(1, 4):  # saving default rows
            for x in range(1, self.sheet_obj.max_column+1):

                eval("R"+str(v)).append((self.sheet_obj.cell(row=v, column=x).value))

        i = 4  # First row with results
        NewRowsStarting = 1  # For numbering combinations

        for x in self.selectedRowList:
            exec("%s  = []" % ("R" + str(i)))  # creating lists dynamically
            for q in range(1, self.sheet_obj.max_column+1):
                if q == 1:
                    eval("R"+str(i)).append(NewRowsStarting)
                else:
                    # storing values of selected row lists
                    eval(
                        "R"+str(i)).append((self.sheet_obj.cell(row=x, column=q).value).strip())
            i += 1
            NewRowsStarting += 1

        # printing results in screen

        # total rows = no. of variables + 3 default rows + 1
        # for display in range(1, len(self.selectedRowList)+4):

        #     print("\t")

        #     for elements in eval("R"+str(display)):
        #         if elements != None:
        #             print(elements, end="   ")

        # creating new workbook
        OutputWb = openpyxl.Workbook()

        OutSheet = OutputWb.active

        OutSheet.title = "OutputFile"

        for addingRows in range(1, len(self.selectedRowList)+4):
            OutSheet.append(eval("R"+str(addingRows)))
        try:
            outputLoc = outputFile
            # if only string given then, file is saved at the location where project is present.
            OutputWb.save(outputLoc)
            return self.resultChanged.emit("success")
            # return "success"
        except:
            self.resultChanged.emit("name")

    def converted(self, a):
        if a == "T":
            return 1
        else:
            return 0

    def Reverse(self, a):
        if a == "T":
            return 0
        elif a == "F":
            return 1
        elif a == 0:
            return 1
        elif a == 1:
            return 0
        else:
            return self.resultChanged.emit("fail")
            # return "fail"
            # print("CELL MUST ONLY CONTAIN 'T' OR 'F'")
            # print("CELL MUST NOT CONTAIN ANY BACKGROUND OR ANY WHITESPACES")

    def makeFinalListFalse(self, list, w):
        tempList = []
        if len(eval(list)) != 0:
            if len(self.selectedRowList) == 0:
                for z in eval(list):
                    count = 1
                    if (z in self.selectedRowList):

                        return None
                    else:
                        for k in range(w+1, self.No_of_Variables + 1):

                            if z in eval("self.VF"+str(k)):
                                count += 1
                        # Used count to get the row with max use in table
                        tempList.append(count)

                return eval(list)[tempList.index(max(tempList))]
            else:
                templist = []
                for o in range(2, self.No_of_Variables+2):
                    exec("%s  = %d" % (
                        "V"+str(o-1), self.converted((self.sheet_obj.cell(row=self.selectedRowList[0], column=o).value).strip())))
                exec("%s  = %d" % ("V"+str(w), self.Reverse(eval("V"+str(w)))))

                for k in eval(list):
                    count = 0
                    for t in range(2, self.No_of_Variables+2):

                        if (self.converted((self.sheet_obj.cell(row=k, column=t).value).strip()) == eval("V"+str(t-1))):
                            count += 1

                    templist.append(count)

                return eval(list)[templist.index(max(templist))]
        else:
            return None

    def makeFinalListTrue(self, check, w, ):

        for o in range(2, self.No_of_Variables+2):

            exec("%s  = %d" % ("V"+str(o-1),
                               self.converted((self.sheet_obj.cell(row=check, column=o).value).strip())))
        exec("%s  = %d" % ("V"+str(w), self.Reverse(eval("V"+str(w)))))

        for l in range(4, self.No_of_combination+4):
            count = 0
            for t in range(2, self.No_of_Variables+2):

                if (self.converted((self.sheet_obj.cell(row=l, column=t).value).strip()) == eval("V"+str(t-1))):
                    count += 1

        # return eval(list)[templist.index(max(templist))]
            if (count == self.No_of_Variables):
                if l not in self.selectedRowList:
                    return self.selectedRowList.append(l)


class AppDemo(QMainWindow):
    def __init__(self):
        super().__init__()
        uic.loadUi('MCDC_UI.ui', self)
        self.showMaximized()

        self.txtFiles = []
        self.excelFiles = []
        self.result = ""
        self.plusTT.clicked.connect(self.openFolderTT)
        self.plusMCDC.clicked.connect(self.openFolderMCDC)
        self.TTlist.itemClicked.connect(self.TTitemClicked)
        self.genTT.clicked.connect(self.genTT_function)
        self.genMCDC.clicked.connect(self.genMCDC_function)

        self.thread = QtCore.QThread(self)
        self.thread.start()

        self.mcdc_program = MCDC_PROGRAM()
        self.mcdc_program.moveToThread(self.thread)
        self.mcdc_program.maximumChanged.connect(self.progressBar.setMaximum)
        self.mcdc_program.progressChanged.connect(self.progressBar.setValue)

        self.mcdc_program.resultChanged.connect(self.resultPublished)

    def openFolderTT(self):
        self.file = str(QFileDialog.getExistingDirectory(
            self, "Select Directory"))
        self.txtFiles = glob.glob(self.file+"/*.txt")
        self.TTlist.clear()
        self.lenOfDir = len(self.file)
        tempList = []
        for a in self.txtFiles:
            tempList.append(a[self.lenOfDir+1:len(a)])
        self.TTlist.addItems(tempList)

    def openFolderMCDC(self):
        self.file = str(QFileDialog.getExistingDirectory(
            self, "Select Directory"))
        self.excelFiles = glob.glob(self.file+"/*TT.xlsx")
        self.MCDClist.clear()
        self.lenOfDir_MCDC = len(self.file)
        tempList = []
        for a in self.excelFiles:
            tempList.append(a[self.lenOfDir_MCDC+1:len(a)])
        self.MCDClist.addItems(tempList)

    def TTitemClicked(self):
        selectedItemIndex = self.TTlist.currentRow()
        selectedFile = self.txtFiles[selectedItemIndex]
        if selectedFile:

            with open(selectedFile) as fr:
                self.expression.setText(fr.read())

    def genTT_function(self):
        calc = truthTableGenerator.Calculating()
        # msg =QErrorMessage()
        msg = QMessageBox()

        if not self.txtFiles and self.expression.text() == "":
            msg.setText(
                "No folder selected or the selected folder donot contain any text files.")
            msg.exec_()

        elif self.TTlist.currentItem() == None and self.expression.text() == "":
            for i in self.txtFiles:
                with open(i) as fr:
                    txt = fr.read()

                result = calc.calculation_truthtb(txt, i[:-4]+"_TT")

                if result == "name":
                    msg.setIcon(QMessageBox.Warning)
                    msg.setWindowTitle("File in use")
                    msg.setText(
                        f"Please close the {i[self.lenOfDir +1:len(i)-4]} excel file.")
                    msg.exec_()

                elif result == "fail":
                    msg.setText("Please check the input file !!")
                    msg.exec_()

            if result == "success":
                msg.setText(
                    "Truth Table of all files saved successfully. (^_^)")
                msg.exec_()

        elif self.expression.text() == '' or len(self.expression.text().strip()) == 0:
            msg.setText("Selected file and dialog box is empty !! ")
            msg.exec_()

        elif self.expression.text() != "" and self.TTlist.currentItem() != None:
            try:
                result = calc.calculation_truthtb(
                    self.expression.text(), self.txtFiles[self.TTlist.currentRow()][:-4]+"_TT")
                if result == "success":
                    msg.setText("File successfully saved (^_^)")
                    msg.exec_()
                    # QMessageBox.about("success","File successfully saved (^_^)")

                elif result == "fail":
                    msg.setText("Please check the input file !!")
                    msg.exec_()

                elif result == "name":
                    msg.setIcon(QMessageBox.Warning)
                    msg.setWindowTitle("File in use")
                    msg.setText("Please close the DialogBox_TT.xlsx file.")
                    msg.exec_()
            except:
                msg.setText(
                    "Sorry, some error occured. Please check the input file.")
                msg.exec_()

        else:
            try:
                result = calc.calculation_truthtb(
                    self.expression.text(), self.file+'/DialogBox_TT')
                if result == "success":
                    msg.setText("File successfully saved (^_^)")
                    msg.exec_()
                    # QMessageBox.about("success","File successfully saved (^_^)")

                elif result == "fail":
                    msg.setText("Please check the input file !!")
                    msg.exec_()

                elif result == "name":
                    msg.setIcon(QMessageBox.Warning)
                    msg.setWindowTitle("File in use")
                    msg.setText("Please close the DialogBox_TT.xlsx file.")
                    msg.exec_()
            except:
                msg.setText(
                    "Sorry, some error occured. Please check the input file.")
                msg.exec_()

    @QtCore.pyqtSlot(str)
    def resultPublished(self, result):
        msg = QMessageBox()
        self.result = result
        if self.result == "success":
            msg.setText("File successfully saved (^_^)")
            msg.exec_()
            self.thread.quit()

            # QMessageBox.about("success","File successfully saved (^_^)")

        elif self.result == "fail":
            self.thread.quit()
            msg.setText("Please check the input file !!")
            msg.exec_()
        elif self.result == "name":
            self.thread.quit()
            msg.setIcon(QMessageBox.Warning)
            msg.setWindowTitle("File in use")
            msg.setText("Please close the excel files.")
            msg.exec_()
        elif self.result == "expression":
            self.thread.quit()
            msg.setText("Please check the expression of input file.")
            msg.exec_()
        else:
            self.thread.quit()
            msg.setText("Sorry, something went wrong.")
            msg.exec_()

    def genMCDC_function(self):
        self.thread.start()
        # cal = mcdcProgram.MCDC_class()
        # msg =QErrorMessage()
        msg = QMessageBox()

        if not self.excelFiles:
            msg.setText(
                "No folder selected or the selected folder donot contain any Truth Table files.")
            msg.exec_()

        elif self.MCDClist.currentItem() == None:
            msg.setText("Please select any file !!")
            msg.exec_()
            # for i in self.excelFiles:
            #     # with open(i) as fr:
            #     #     txt = fr.read()

            #     result = cal.CreateMCDC(i, i[:-5]+"_MCDC.xlsx")

            #     if result == "name":
            #         msg.setText(
            #             f"Please close the {i[self.lenOfDir +1:len(i)-4]} excel file.")
            #         msg.exec_()

            #     elif result == "fail":
            #         msg.setText("Please check the input file !!")
            #         msg.exec_()

            # if result == "success":
            #     msg.setText("All files saved successfully. (^_^)")
            #     msg.exec_()

        else:
            wrapper = partial(self.mcdc_program.CreateMCDC, self.excelFiles[self.MCDClist.currentRow(
            )], self.excelFiles[self.MCDClist.currentRow()][:-7]+"MCDC.xlsx")
            QtCore.QTimer.singleShot(0, wrapper)
            # result = cal.CreateMCDC(
            # self.excelFiles[self.MCDClist.currentRow()] , self.excelFiles[self.MCDClist.currentRow()][:-7]+"MCDC.xlsx")

            # if self.result == "success":
            #     msg.setText("File successfully saved (^_^)")
            #     msg.exec_()
            #     # QMessageBox.about("success","File successfully saved (^_^)")

            # elif self.result ==  "fail":
            #     msg.setText("Please check the input file !!")
            #     msg.exec_()

            # elif self.result == "name":
            #     msg.setIcon(QMessageBox.Warning)
            #     msg.setWindowTitle("File in use")
            #     msg.setText("Please close the excel files.")
            #     msg.exec_()
            # elif self.result == "expression":
            #     msg.setText("Please check the expression of input file.")
            #     msg.exec_()
            # else:
            #     msg.setText("Sorry, something went wrong.")
            #     msg.exec_()

    # def startProgressBar(self):
    #     self.thread = MyThread()
    #     self.thread.change_value.connect(self.setProgressVal)
    #     self.thread.start()

    # def setProgressVal(self,val):
    #     self.progressBar.setValue(val)

# class MyThread(QThread):
#     change_value = pyqtSignal(int)

#     def run(self):
#         self.cnt = 0
#         while self.cnt < 100:
#             self.cnt+=1
#             time.sleep(0.5)
#             self.change_value.emit(self.cnt)

#     def genMCDC_function(self):
#         AppDemo().startProgressBar()
#         cal = mcdcProgram.MCDC_class()
#         # msg =QErrorMessage()
#         msg = QMessageBox()

#         if not self.excelFiles:
#             msg.setText(
#                 "No folder selected or the selected folder donot contain any Truth Table files.")
#             msg.exec_()

#         elif self.MCDClist.currentItem() == None:
#             msg.setText("Please select any file !!")
#             msg.exec_()
#             # for i in self.excelFiles:
#             #     # with open(i) as fr:
#             #     #     txt = fr.read()

#             #     result = cal.CreateMCDC(i, i[:-5]+"_MCDC.xlsx")

#             #     if result == "name":
#             #         msg.setText(
#             #             f"Please close the {i[self.lenOfDir +1:len(i)-4]} excel file.")
#             #         msg.exec_()

#             #     elif result == "fail":
#             #         msg.setText("Please check the input file !!")
#             #         msg.exec_()

#             # if result == "success":
#             #     msg.setText("All files saved successfully. (^_^)")
#             #     msg.exec_()


#         else:

#             result = cal.CreateMCDC(
#                 self.excelFiles[self.MCDClist.currentRow()] , self.excelFiles[self.MCDClist.currentRow()][:-7]+"MCDC.xlsx")
#             self.cnt == 99
#             if result == "success":
#                 msg.setText("File successfully saved (^_^)")
#                 msg.exec_()
#                 # QMessageBox.about("success","File successfully saved (^_^)")

#             elif result == "fail":
#                 msg.setText("Please check the input file !!")
#                 msg.exec_()

#             elif result == "name":
#                 msg.setIcon(QMessageBox.Warning)
#                 msg.setWindowTitle("File in use")
#                 msg.setText("Please close the excel files.")
#                 msg.exec_()
#             elif result == "expression":
#                 msg.setText("Please check the expression of input file.")
#                 msg.exec_()
#             else:
#                 msg.setText("Sorry, something went wrong.")
#                 msg.exec_()

if __name__ == '__main__':
    # app= QApplication(sys.argv)
    app = QApplication(sys.argv)

    demo = AppDemo()
    demo.show()

    try:
        sys.exit(app.exec_())
    except SystemExit:
        print('Closing window........')
